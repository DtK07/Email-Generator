import openpyxl
from openpyxl import workbook, load_workbook
wb = openpyxl.load_workbook("Emails.xlsx")
sheet1 = wb['Email']
wb.create_sheet("Email_guesses")
sh = wb["Email_guesses"]
sh['A1'].value = "Name"
sh['B1'].value = "First guess"
sh['C1'].value = "Second guess"
sh['D1'].value = "Third guess"
sh['E1'].value = "Fourth guess"
sh['F1'].value = "Fifth guess"
sh['G1'].value = "Sixth guess"
row = sheet1.max_row
column = sheet1.max_column
def mail_guesser(first_name, domain, last_name = None):
    guess = []
    if last_name == None:
        first_guess = first_name.lower() + "@" + domain
        guess.append((first_guess))
    else:
        first_guess = first_name.lower() + "@" + domain
        second_guess = last_name.lower() + "@" + domain
        third_guess = first_name.lower() + "." + last_name.lower() + "@" + domain
        fourth_guess = first_name[0].lower() + last_name.lower() + "@" + domain
        fifth_guess = first_name.lower() + last_name[0].lower() + "@" + domain
        sixth_guess = first_name.lower() + last_name.lower() + "@" + domain
        guess.append((first_guess, second_guess, third_guess, fourth_guess, fifth_guess, sixth_guess))
    return guess
def name_split (name):
    list =[]
    name_list = name.split(" ")
    #check length of a name and cut short the name to first and last name irrespective of the total length
    if len(name_list) == None:
        pass
    elif len(name_list) == 1:
        first_name = name_list[0]
        last_name = None
        list.append((first_name, last_name))
    elif len(name_list) == 2:
        first_name = name_list[0]
        last_name = name_list[1]
        list.append((first_name, last_name))
    elif len(name_list) == 3:
        first_name = name_list[0]
        last_name = name_list[2]
        list.append((first_name, last_name))
    elif len(name_list) == 4:
        first_name = name_list[0]
        last_name = name_list[3]
        list.append((first_name, last_name))
    return list
i = 2
while i < row+1:
    j = 1
    name = sheet1.cell(i,j).value
    j = 2
    domain = sheet1.cell(i,j).value
    if domain == None:
        j = 3
        website = sheet1.cell(i,j).value
        if website == None:
            pass
        else:
            domain = website.split("/")[2]
            if "www" in domain:
                domain = domain[4:]
    name_func = name_split(name)
    #print(name_func)
    j = 1
    for name in name_func:
        first_name, last_name = name
        if last_name == None:
            sh.cell(i,j).value = first_name
        else:
            sh.cell(i,j).value = first_name + " " + last_name
    mail = mail_guesser(first_name, domain, last_name)
    print(mail)
    j = 2
    if last_name == None:
        sh.cell(i,j).value = mail[0]
    else:
        for id in mail:
            for mail in id:
                sh.cell(i,j).value = mail
                print(mail)
                j+=1
    i += 1
wb.save("Emails.xlsx")









